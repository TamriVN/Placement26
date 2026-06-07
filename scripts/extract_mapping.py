#!/usr/bin/env python3
"""Extract VNVC form variables and FHIR mapping from docs/*.xlsx into JSON."""

from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path

import openpyxl

from fhir_links import enrich_fhir_links, is_fhir_resource

ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
OUTPUT = ROOT / "output"

# Master dataset column → form sheet name (from Mapped_dataset.xlsx headers)
FORM_COLUMN_TO_SHEET = {
    "form_item_form_1": "Form_1",
    "form_item_form_2": "Form_2",
    "form_item_form_3": "Form_3",
    "form_item_form_4": "Form_4",
}


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_header_row(ws, markers: tuple[str, ...]) -> tuple[int, list[str]]:
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [_clean(c) for c in row]
        if any(m in cells for m in markers):
            return idx, cells
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return 1, [_clean(c) for c in first]


def _is_unmapped(resource: str) -> bool:
    lowered = resource.lower()
    return not resource or "does not map" in lowered


def parse_form_titles(path: Path) -> dict[str, str]:
    """Read form titles from cell A1 of each sheet (docs source)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    titles: dict[str, str] = {}
    for sheet_name in wb.sheetnames:
        raw = _clean(wb[sheet_name]["A1"].value)
        # e.g. "Variable Extraction: Form 1 (AEFI_Report) on Common Post Vaccination Reactions"
        title = re.sub(r"^Variable Extraction:\s*", "", raw).strip()
        titles[sheet_name.strip()] = title or raw
    return titles


def build_form_labels(form_titles: dict[str, str]) -> dict[str, str]:
    """Map Master dataset column keys to form titles from docs."""
    labels: dict[str, str] = {}
    for column, sheet_key in FORM_COLUMN_TO_SHEET.items():
        labels[column] = form_titles.get(sheet_key, sheet_key)
    return labels


def parse_mapped_dataset(path: Path, form_labels: dict[str, str]) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Master dataset"]
    header_row, headers = _find_header_row(ws, ("Parent header", "form_item_form_1", "concept"))
    records: list[dict] = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        data = {_clean(headers[i]): _clean(row[i]) for i in range(len(headers)) if i < len(row)}
        if not (data.get("concept") or data.get("fhir_resource")):
            continue

        forms = []
        for key, label in form_labels.items():
            field = data.get(key, "")
            if field:
                forms.append({"form_key": key, "form_label": label, "field": field})

        fhir_resource = data.get("fhir_resource", "")
        fhir_element = data.get("fhir_element", "")
        records.append(
            {
                "concept": data.get("concept", ""),
                "parent_header": data.get("Parent header", ""),
                "vnvc_data_type": data.get("vnvc_data_type", ""),
                "forms": forms,
                "fhir": {
                    "resource": fhir_resource,
                    "element": fhir_element,
                    "definition": data.get("fhir_definition", ""),
                    "data_type": data.get("fhir_data_type", ""),
                    "links": enrich_fhir_links(fhir_resource, fhir_element),
                },
                "terminology": data.get("terminology", ""),
                "notes": data.get("Notes", ""),
                "unmapped": _is_unmapped(fhir_resource),
            }
        )
    return records


def parse_form_variables(path: Path) -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    result: dict[str, list[dict]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, headers = _find_header_row(ws, ("Variable ID", "Variable ID "))
        normalized_headers = [
            h.replace("Variable ID ", "variable_id").replace("Variable ID", "variable_id") for h in headers
        ]

        rows: list[dict] = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            data = {
                normalized_headers[i]: _clean(row[i])
                for i in range(len(normalized_headers))
                if i < len(row)
            }
            vid = data.get("variable_id", "")
            if vid and re.match(r"^F\d", vid):
                rows.append(data)

        key = sheet_name.strip()
        result[key] = rows
    return result


def parse_data_dictionary(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_row, headers = _find_header_row(ws, ("Form Source", "Data element (English translated)"))
    rows: list[dict] = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        data = {_clean(headers[i]): _clean(row[i]) for i in range(len(headers)) if i < len(row)}
        if data.get("Data element (English translated)") or data.get("Form Source"):
            rows.append(data)
    return rows


def parse_fhir_resource_json(path: Path) -> dict[str, str]:
    """JSON for Resources sheet from Mapped_dataset.xlsx."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["JSON for Resources"]
    result: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        name = _clean(row[0]) if row else ""
        body = _clean(row[1]) if row and len(row) > 1 else ""
        if name and body:
            result[name] = body
    return result


def parse_ernest_template(path: Path) -> list[dict]:
    """Ernest template sheet from Mapped_dataset.xlsx."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Ernest template"]
    header_row, headers = _find_header_row(ws, ("form_item", "fhir_resource"))
    rows: list[dict] = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        data = {_clean(headers[i]): _clean(row[i]) for i in range(len(headers)) if i < len(row)}
        if data.get("form_item") or data.get("fhir_resource"):
            resource = data.get("fhir_resource", "")
            element = data.get("fhir_element", "")
            data["links"] = enrich_fhir_links(resource, element)
            rows.append(data)
    return rows


def build_form_fhir_matrix(records: list[dict]) -> dict:
    """Form → FHIR resource counts derived only from Mapped_dataset rows."""
    matrix: dict[str, dict[str, int]] = {}
    for rec in records:
        resource = "Unmapped / Gap" if rec["unmapped"] else rec["fhir"]["resource"]
        if not rec["forms"]:
            matrix.setdefault("(no form column)", {})
            matrix["(no form column)"][resource] = matrix["(no form column)"].get(resource, 0) + 1
            continue
        for form in rec["forms"]:
            label = form["form_label"]
            matrix.setdefault(label, {})
            matrix[label][resource] = matrix[label].get(resource, 0) + 1
    return matrix


def build_stats(
    mapping_records: list[dict],
    forms_data: dict,
    form_titles: dict[str, str],
    form_labels: dict[str, str],
    dictionary_rows: list[dict],
) -> dict:
    return {
        "total_concepts": len(mapping_records),
        "mapped": sum(1 for r in mapping_records if not r["unmapped"]),
        "unmapped": sum(1 for r in mapping_records if r["unmapped"]),
        "by_resource": dict(
            Counter(r["fhir"]["resource"] for r in mapping_records if not r["unmapped"])
        ),
        "by_form": {
            label: sum(1 for r in mapping_records if any(f["form_key"] == key for f in r["forms"]))
            for key, label in form_labels.items()
        },
        "form_variable_counts": {form_titles.get(k, k): len(v) for k, v in forms_data.items()},
        "data_dictionary_forms": sorted(
            {r.get("Form Source", "") for r in dictionary_rows if r.get("Form Source")}
        ),
    }


def main() -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)

    forms_path = DOCS / "Initial_Extraction_of_Variables_from_VNVC_forms.xlsx"
    mapped_path = DOCS / "Mapped_dataset.xlsx"
    dict_path = DOCS / "VNVC_Data_Dictionary.xlsx"

    form_titles = parse_form_titles(forms_path)
    form_labels = build_form_labels(form_titles)

    mapping_records = parse_mapped_dataset(mapped_path, form_labels)
    forms_data = parse_form_variables(forms_path)
    dictionary_rows = parse_data_dictionary(dict_path)
    fhir_resource_json = parse_fhir_resource_json(mapped_path)
    ernest_template = parse_ernest_template(mapped_path)

    used_resources = sorted(
        {r["fhir"]["resource"] for r in mapping_records if not r["unmapped"] and is_fhir_resource(r["fhir"]["resource"])}
    )

    model = {
        "source_files": [
            "docs/Mapped_dataset.xlsx",
            "docs/Initial_Extraction_of_Variables_from_VNVC_forms.xlsx",
            "docs/VNVC_Data_Dictionary.xlsx",
        ],
        "form_titles": form_titles,
        "form_labels": form_labels,
        "fhir_spec": {
            "version": "R5",
            "base_url": "https://hl7.org/fhir/R5",
            "resources": {res: enrich_fhir_links(res, "") for res in used_resources},
        },
        "mapping_records": mapping_records,
        "form_fhir_matrix": build_form_fhir_matrix(mapping_records),
        "forms_data": forms_data,
        "data_dictionary": dictionary_rows,
        "data_dictionary_count": len(dictionary_rows),
        "fhir_resource_json": fhir_resource_json,
        "ernest_template": ernest_template,
        "stats": build_stats(mapping_records, forms_data, form_titles, form_labels, dictionary_rows),
    }

    out_path = OUTPUT / "vnvc_fhir_model.json"
    out_path.write_text(json.dumps(model, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {out_path}")
    print(json.dumps(model["stats"], indent=2))


if __name__ == "__main__":
    main()
